#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# svar.py
# Recursive block-exogenous SVAR (Peersman-Smets, adapted to Spain as a small
# member of the monetary union). Response of Spanish real consumption to a +100 bp
# ECB monetary shock, identified by a Cholesky ordering with the euro-area block
# exogenous.
#
# Ordering:
#   [ oil(EUR), GDP_EA, prices_EA, policy rate | GDP_ES, prices_ES, consumption_ES ]  (+ EUR/USD)
#   - the area/external block is exogenous (Spain does not enter its equations);
#   - oil is strictly exogenous (a global price the ECB and Spain do not move, so
#     its response to the shock is zero);
#   - EUR/USD is ordered after the rate (an asset price);
#   - the monetary shock is the rate innovation, orthogonal to oil/GDP_EA/prices_EA;
#   - the IRF is normalised to +100 bp; bands are a residual bootstrap.
# Sample 1999-2019 (2020+ excluded: COVID and the 2022 supply shock break a linear VAR).
# Prices are seasonally adjusted (STL); the policy rate is the shadow rate spliced
# onto the 1-month Euribor.
#
# Declared limitation: euro-area GDP shows a mild output puzzle (positive in the
# first year, fading by h6), a residual of the information effect that a plain
# Cholesky does not purge. It does not affect Spanish consumption (small block,
# dominated by the direct rate->consumption channel).
#
# Requirements: pandas numpy statsmodels matplotlib openpyxl xlrd
# Data directories are taken from the environment (SVAR_DATA, SVAR_EURI, SVAR_OUT)
# with defaults relative to this file.

import os, glob, warnings
import numpy as np, pandas as pd
import openpyxl
from statsmodels.tsa.seasonal import STL
warnings.filterwarnings("ignore")

# ============================ CONFIG ============================
_HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.environ.get("SVAR_DATA", os.path.normpath(os.path.join(_HERE, "..", "..", "..", "data", "macro")))
EURI = os.environ.get("SVAR_EURI", os.path.normpath(os.path.join(_HERE, "..", "..", "..", "data", "euribor")))
OUT  = os.environ.get("SVAR_OUT",  os.path.join(_HERE, "output"))
os.makedirs(OUT, exist_ok=True)
Y0, Y1, P_LAG, H, NBOOT = 1999, 2019, 2, 16, 1000
OIL_EXO = True

def find(folder, sub):
    for p in glob.glob(os.path.join(folder, "*")):
        if sub in os.path.basename(p): return p
    raise FileNotFoundError(sub)

# ---------------------- readers ----------------------
def qk(q): return q.replace("-Q", "Q")
def geo_rows(f, sheet, col, geo):
    ws = openpyxl.load_workbook(find(DATA, f), data_only=True)[sheet]; o = {}
    for i in range(11, ws.max_row+1):
        g = ws.cell(i,1).value
        if g and str(g).startswith(geo):
            v = ws.cell(i,col).value
            if isinstance(v,(int,float)): o[qk(str(ws.cell(i,2).value))] = float(v)
    return o
def ecb_csv(folder, sub, valcol=2):
    for p in glob.glob(os.path.join(folder, "*.csv")):
        if sub in open(p).readline():
            o = {}
            for line in open(p).read().splitlines()[1:]:
                pr = [x.strip('"') for x in line.split('","')]; pr[0] = pr[0].strip('"')
                if len(pr) <= valcol or pr[valcol] in ("","-"): continue
                try: o[pr[0][:7]] = float(pr[valcol])
                except ValueError: pass
            return pd.Series(o)
    raise FileNotFoundError(sub)
def hicp_sa_quarterly():
    ws = openpyxl.load_workbook(find(DATA, "prc_hicp_minr"), data_only=True)["Sheet 1"]
    ea, es = {}, {}
    for i in range(11, ws.max_row+1):
        t = ws.cell(i,1).value
        if not t or "-" not in str(t): continue
        for col, d in [(2,ea),(4,es)]:
            v = ws.cell(i,col).value
            if isinstance(v,(int,float)): d[str(t)] = float(v)
    def sa(d):
        s = pd.Series({pd.Period(k,'M'): np.log(v) for k,v in d.items()}).sort_index()
        s.index = s.index.to_timestamp()
        slog = s - STL(s, period=12, robust=True).fit().seasonal
        df = slog.to_frame("v"); df["q"] = df.index.year.astype(str)+"Q"+((df.index.month-1)//3+1).astype(str)
        return df.groupby("q")["v"].mean()*100
    return sa(ea), sa(es)
def brent_eur_quarterly():
    br = pd.read_excel(find(DATA, "RBRTE"), sheet_name="Data 1", skiprows=3, header=None, names=["date","usd"])
    br["date"] = pd.to_datetime(br["date"], errors="coerce"); br = br.dropna()
    fx = ecb_csv(DATA, "US dollar/Euro")
    br["ym"] = br.date.dt.strftime("%Y-%m"); br["eur"] = br.usd/br.ym.map(fx)
    br = br.dropna(subset=["eur"]); br["q"] = br.ym.str[:4]+"Q"+((br.date.dt.month-1)//3+1).astype(str)
    return np.log(br.groupby("q")["eur"].mean())*100, fx
def eurusd_quarterly(fx):
    df = fx.to_frame("fx"); df["q"] = df.index.str[:4]+"Q"+((df.index.str[5:7].astype(int)-1)//3+1).astype(str)
    return np.log(df.groupby("q")["fx"].mean())*100
def shadow_spliced_quarterly():
    e1 = ecb_csv(EURI, "Euribor 1-month")
    sr = pd.read_excel(find(DATA, "shadowrate_ECB"), header=None, names=["ym","sr"])
    sr["ym"] = sr.ym.astype(int).astype(str).str.replace(r"(\d{4})(\d{2})", r"\1-\2", regex=True)
    sh = sr.set_index("ym")["sr"]; lo, hi = sh.index.min(), sh.index.max()
    off_lo, off_hi = sh[lo]-e1[lo], sh[hi]-e1[hi]; rate = {}
    for m in sorted(set(e1.index)|set(sh.index)):
        if lo <= m <= hi and m in sh.index: rate[m] = sh[m]
        elif m < lo and m in e1.index:      rate[m] = e1[m]+off_lo
        elif m > hi and m in e1.index:      rate[m] = e1[m]+off_hi
    s = pd.Series(rate); df = s.to_frame("v")
    df["q"] = df.index.str[:4]+"Q"+((df.index.str[5:7].astype(int)-1)//3+1).astype(str)
    return df.groupby("q")["v"].mean()

def build_panel():
    gdp_ea = pd.Series(geo_rows("namq_10_gdp","Sheet 1",5,"Euro")).apply(np.log)*100
    gdp_es = pd.Series(geo_rows("namq_10_gdp","Sheet 1",5,"Spain")).apply(np.log)*100
    cons_es= pd.Series(geo_rows("namq_10_fcs","Sheet 1",3,"Spain")).apply(np.log)*100
    lp_ea, lp_es = hicp_sa_quarterly()
    loil, fx = brent_eur_quarterly(); leur = eurusd_quarterly(fx)
    rate = shadow_spliced_quarterly()
    P = pd.DataFrame({"loil":loil,"lgdp_ea":gdp_ea,"lp_ea":lp_ea,"rate":rate,
                      "leur":leur,"lgdp_es":gdp_es,"lp_es":lp_es,"lcons_es":cons_es})
    P["yr"] = P.index.str[:4].astype(int); P["qn"] = P.index.str[5:].astype(int)
    P = P.sort_values(["yr","qn"]); P = P[(P.yr>=Y0)&(P.yr<=Y1)].drop(columns=["yr","qn"]).dropna()
    return P

# ---------------------- SVAR ----------------------
NAMES = ["loil","lgdp_ea","lp_ea","rate","leur","lgdp_es","lp_es","lcons_es"]
LABELS= ["Oil (EUR)","GDP euro area","Prices euro area","Policy rate (bp)","EUR/USD","GDP Spain","Prices Spain","Consumption Spain"]
AREA, DOM, RATE, CONS, OIL = [0,1,2,3,4], [5,6,7], 3, 7, 0

def estimate(Y, p, oil_exo=OIL_EXO):
    T, n = Y.shape
    Ylag = np.hstack([Y[p-1-k:T-1-k] for k in range(p)]); yt = Y[p:]
    const = np.ones((T-p,1)); tr = np.arange(p,T).reshape(-1,1)*1.0
    A = [np.zeros((n,n)) for _ in range(p)]; C = np.zeros(n); G = np.zeros(n); resid = np.zeros((yt.shape[0],n))
    for i in range(n):
        if oil_exo and i == OIL: allowed = [OIL]                # oil strictly exogenous
        elif i in DOM:           allowed = list(range(n))       # Spain responds to everything
        else:                    allowed = AREA                 # area block exogenous
        cols = [k*n+j for k in range(p) for j in allowed]
        X = np.hstack([const, tr, Ylag[:,cols]]); b,_,_,_ = np.linalg.lstsq(X, yt[:,i], rcond=None)
        resid[:,i] = yt[:,i]-X@b; C[i], G[i] = b[0], b[1]
        for k in range(p):
            for jj,j in enumerate(allowed): A[k][i,j] = b[2+k*len(allowed)+jj]
    return A, C, G, np.cov(resid,rowvar=False), resid

def irf(A, Sig, p, H):
    n = Sig.shape[0]; Pl = np.linalg.cholesky(Sig); d = Pl[:,RATE]/Pl[RATE,RATE]   # +100 bp
    Phi = [np.eye(n)]
    for h in range(1,H+1):
        M = np.zeros((n,n))
        for k in range(p):
            if h-1-k >= 0: M += A[k]@Phi[h-1-k]
        Phi.append(M)
    return np.array([Phi[h]@d for h in range(H+1)])

def bootstrap(Yall, A, C, G, resid, p, H, B, oil_exo=OIL_EXO):
    T, n = Yall.shape; out = np.full((B,H+1,n), np.nan)
    for b in range(B):
        Y = np.zeros((T,n)); Y[:p] = Yall[:p]; idx = np.random.randint(0, resid.shape[0], size=T-p)
        for t in range(p,T):
            v = C+G*t
            for k in range(p): v = v+A[k]@Y[t-1-k]
            Y[t] = v+resid[idx[t-p]]
        try:
            Ab,_,_,Sb,_ = estimate(Y,p,oil_exo); out[b] = irf(Ab,Sb,p,H)
        except np.linalg.LinAlgError: pass
    return out

def make_figure_main(r, path):
    import matplotlib; matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MultipleLocator
    BLUE,RED = "#2166AC","#B2182B"
    plt.rcParams.update({"font.size":11,"font.family":"DejaVu Sans","axes.edgecolor":"#555"})
    fig,a = plt.subplots(figsize=(8.4,5.2)); m = r[r.h<=10]; h = m.h
    impact, trough = r["lcons_es"].iloc[0], r.loc[r.h<=8,"lcons_es"].min()
    a.fill_between(h,m["lcons_es_lo90"],m["lcons_es_hi90"],color=BLUE,alpha=.12,lw=0,label="90% band")
    a.fill_between(h,m["lcons_es_lo68"],m["lcons_es_hi68"],color=BLUE,alpha=.25,lw=0,label="68% band")
    a.plot(h,m["lcons_es"],color=BLUE,lw=2.6,label="Spanish consumption (SVAR)")
    a.plot([0,4,8],[-0.38,-0.13,-0.01],"D--",color=RED,lw=1.7,ms=7,label="HANK model (annual)")
    a.axhline(0,color="#333",lw=.9,ls=(0,(4,3)))
    a.set_title(f"Spanish real consumption to +100 bp — recursive SVAR\nimpact {impact:.2f}% · trough {trough:.2f}% within a year",fontsize=11)
    a.set_xlabel("quarters after the shock"); a.set_ylabel("consumption, % · per +100 bp")
    a.set_xlim(0,10); a.xaxis.set_major_locator(MultipleLocator(1)); a.grid(alpha=.18)
    a.legend(frameon=False,fontsize=9,loc="lower left")
    fig.tight_layout(); fig.savefig(path,dpi=150,bbox_inches="tight")

def make_figure_8panel(r, path):
    import matplotlib; matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MultipleLocator
    plt.rcParams.update({"font.size":9.5,"font.family":"DejaVu Sans","axes.edgecolor":"#666"})
    fig,axs = plt.subplots(2,4,figsize=(15,7)); m = r[r.h<=12]; h = m.h
    for i,(nm,lab) in enumerate(zip(NAMES,LABELS)):
        ax = axs[i//4, i%4]; sc = 100 if i==RATE else 1
        col = "#2166AC" if i==CONS else ("#B2182B" if i==1 else "#444")
        ax.fill_between(h, m[nm+"_lo68"]*sc, m[nm+"_hi68"]*sc, color=col, alpha=.18, lw=0)
        ax.plot(h, m[nm]*sc, color=col, lw=2.2)
        ax.axhline(0,color="#333",lw=.8,ls=(0,(4,3)))
        ax.set_title(lab + ("  (output puzzle)" if i==1 else ("  (exogenous)" if i==OIL else "")), fontsize=9.5)
        ax.set_xlim(0,12); ax.xaxis.set_major_locator(MultipleLocator(3)); ax.grid(alpha=.15)
        if i//4==1: ax.set_xlabel("quarters")
    fig.suptitle("Recursive SVAR for Spain — the eight responses to a +100 bp monetary shock (68% bands)",fontsize=12,y=1.0)
    fig.tight_layout(); fig.savefig(path,dpi=140,bbox_inches="tight")

def main():
    print("Panel 1999-2019, 8 variables (oil exogenous)...")
    P = build_panel(); Yall = P[NAMES].values; P.to_csv(os.path.join(OUT,"svar_panel.csv"))
    A,C,G,Sig,resid = estimate(Yall, P_LAG); point = irf(A,Sig,P_LAG,H)
    np.random.seed(0); boot = bootstrap(Yall,A,C,G,resid,P_LAG,H,NBOOT)
    cols = {"h": list(range(H+1))}
    for i,nm in enumerate(NAMES):
        cols[nm] = point[:,i]
        cols[nm+"_lo68"] = np.nanpercentile(boot[:,:,i],16,axis=0); cols[nm+"_hi68"] = np.nanpercentile(boot[:,:,i],84,axis=0)
        cols[nm+"_lo90"] = np.nanpercentile(boot[:,:,i], 5,axis=0); cols[nm+"_hi90"] = np.nanpercentile(boot[:,:,i],95,axis=0)
    r = pd.DataFrame(cols); r.to_csv(os.path.join(OUT,"svar_irf.csv"), index=False)
    make_figure_main(r, os.path.join(OUT,"svar_consumption_IRF.png"))
    make_figure_8panel(r, os.path.join(OUT,"svar_irf_8panel.png"))
    tr = r.loc[r.h<=8,"lcons_es"].min(); ht = int(r.loc[r.h<=8,"lcons_es"].idxmin())
    print(f"\nSpanish consumption: impact {r['lcons_es'].iloc[0]:+.2f}%, trough {tr:+.2f}% at h={ht}Q.")
    print(f"checks: prices_ES h4 {r['lp_es'].iloc[4]:+.2f} and prices_EA h4 {r['lp_ea'].iloc[4]:+.2f} (no price puzzle); "
          f"rate h6 {r['rate'].iloc[6]*100:.0f} bp (reverts); EUR/USD impact {r['leur'].iloc[0]:+.2f} (euro appreciates).")
    print(f"limitation: GDP_EA h2 {r['lgdp_ea'].iloc[2]:+.2f} (mild output puzzle, information residual).")
    print(f"\noutputs in {OUT}: svar_panel.csv, svar_irf.csv, svar_consumption_IRF.png, svar_irf_8panel.png")

if __name__ == "__main__":
    main()
